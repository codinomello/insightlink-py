"""
reports.py
-----------
Geração de relatórios (PDF e Excel) a partir dos registros importados,
consolidando KPIs, cruzamentos de dados e a listagem completa.
"""
import io
from datetime import datetime

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak,
)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

NAVY = colors.HexColor("#1b3a5c")
ORANGE = colors.HexColor("#f5a623")


def build_pdf_report(summary: dict, analytics: dict, records: list[dict]) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2 * cm, bottomMargin=2 * cm,
                             leftMargin=2 * cm, rightMargin=2 * cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Title"], textColor=NAVY)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], textColor=NAVY, spaceBefore=14, spaceAfter=8)
    normal = styles["Normal"]

    story = [
        Paragraph("InsightLink — Relatório Gerencial", title_style),
        Paragraph(f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}", normal),
        Spacer(1, 16),
    ]

    # KPIs
    story.append(Paragraph("Indicadores Gerais", h2))
    kpi_rows = [
        ["Total de registros", summary.get("total", 0)],
        ["Desafios", summary.get("desafios", 0)],
        ["Projetos", summary.get("projetos", 0)],
        ["Empresas envolvidas", summary.get("empresas", 0)],
        ["Proponentes únicos", summary.get("proponentes", 0)],
        ["Completude média (%)", summary.get("completude_media", 0)],
        ["Palavras em média por registro", summary.get("palavras_media", 0)],
    ]
    t = Table(kpi_rows, colWidths=[9 * cm, 6 * cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.whitesmoke),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("PADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(t)

    # Registros por empresa
    story.append(Paragraph("Registros por Empresa", h2))
    dados_empresa = [["Empresa", "Quantidade"]] + [
        [e["nome"], e["quantidade"]] for e in summary.get("por_empresa", [])
    ]
    t2 = Table(dados_empresa, colWidths=[9 * cm, 6 * cm])
    t2.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ("PADDING", (0, 0), (-1, -1), 5),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.whitesmoke]),
    ]))
    story.append(t2)

    # Cruzamento completude por tipo
    comp_tipo = analytics.get("completude_por_tipo", [])
    if comp_tipo:
        story.append(Paragraph("Completude por Tipo (Desafio x Projeto)", h2))
        dados_tipo = [["Tipo", "Média", "Mínimo", "Máximo", "Qtd."]] + [
            [c["tipo"], c["media"], c["minimo"], c["maximo"], c["quantidade"]]
            for c in comp_tipo
        ]
        t3 = Table(dados_tipo, colWidths=[4 * cm, 3 * cm, 3 * cm, 3 * cm, 2 * cm])
        t3.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), ORANGE),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ("PADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(t3)

    # Correlação
    corr = analytics.get("correlacao_palavras_completude", {})
    if corr:
        story.append(Paragraph("Correlação: Volume de Texto x Completude", h2))
        story.append(Paragraph(
            f"Coeficiente de correlação de Pearson: <b>{corr.get('coeficiente', 0)}</b> "
            "(valores próximos de 1 indicam que registros mais detalhados tendem a "
            "preencher mais campos do formulário).", normal,
        ))

    # Listagem completa
    story.append(PageBreak())
    story.append(Paragraph("Listagem Completa de Registros", h2))
    header = ["Título", "Tipo", "Empresa", "Cargo", "Completude"]
    dados_lista = [header] + [
        [
            (r.get("titulo") or "")[:40],
            r.get("tipo") or "",
            (r.get("empresa") or "")[:25],
            (r.get("cargo") or "")[:25],
            f"{r.get('completude', 0)}%",
        ]
        for r in records
    ]
    t4 = Table(dados_lista, colWidths=[5.5 * cm, 2.2 * cm, 4 * cm, 4 * cm, 2.3 * cm], repeatRows=1)
    t4.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.lightgrey),
        ("PADDING", (0, 0), (-1, -1), 4),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.whitesmoke]),
    ]))
    story.append(t4)

    doc.build(story)
    buffer.seek(0)
    return buffer.read()


def build_xlsx_report(summary: dict, analytics: dict, records: list[dict]) -> bytes:
    wb = Workbook()
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1B3A5C")
    body_font = Font(name="Arial", size=10)

    def write_table(ws, headers, rows, start_row=1):
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for r, row in enumerate(rows, start=start_row + 1):
            for c, value in enumerate(row, start=1):
                cell = ws.cell(row=r, column=c, value=value)
                cell.font = body_font
        for c, h in enumerate(headers, start=1):
            width = max([len(str(h))] + [len(str(row[c - 1])) for row in rows]) if rows else len(str(h))
            ws.column_dimensions[get_column_letter(c)].width = min(width + 4, 45)

    # Aba 1: Resumo (KPIs)
    ws1 = wb.active
    ws1.title = "Resumo"
    write_table(ws1, ["Indicador", "Valor"], [
        ["Total de registros", summary.get("total", 0)],
        ["Desafios", summary.get("desafios", 0)],
        ["Projetos", summary.get("projetos", 0)],
        ["Empresas envolvidas", summary.get("empresas", 0)],
        ["Proponentes únicos", summary.get("proponentes", 0)],
        ["Completude média (%)", summary.get("completude_media", 0)],
        ["Palavras em média por registro", summary.get("palavras_media", 0)],
    ])

    # Aba 2: Por empresa
    ws2 = wb.create_sheet("Por Empresa")
    write_table(ws2, ["Empresa", "Quantidade"],
                [[e["nome"], e["quantidade"]] for e in summary.get("por_empresa", [])])

    # Aba 3: Por cargo
    ws3 = wb.create_sheet("Por Cargo")
    write_table(ws3, ["Cargo", "Quantidade"],
                [[c["nome"], c["quantidade"]] for c in summary.get("por_cargo", [])])

    # Aba 4: Cruzamento Empresa x Cargo
    ws4 = wb.create_sheet("Cruzamento Empresa x Cargo")
    ct = analytics.get("crosstab_empresa_cargo", {})
    write_table(ws4, ["Empresa", "Cargo", "Quantidade"],
                [[m["empresa"], m["cargo"], m["quantidade"]] for m in ct.get("matriz", [])])

    # Aba 5: Completude por tipo
    ws5 = wb.create_sheet("Completude por Tipo")
    write_table(ws5, ["Tipo", "Média", "Mínimo", "Máximo", "Quantidade"],
                [[c["tipo"], c["media"], c["minimo"], c["maximo"], c["quantidade"]]
                 for c in analytics.get("completude_por_tipo", [])])

    # Aba 6: Listagem completa
    ws6 = wb.create_sheet("Registros")
    write_table(ws6, ["Título", "Tipo", "Proponente", "E-mail", "Telefone", "Empresa", "Cargo", "Completude"],
                [[r.get("titulo"), r.get("tipo"), r.get("proponente"), r.get("email"),
                  r.get("telefone"), r.get("empresa"), r.get("cargo"), r.get("completude")]
                 for r in records])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()